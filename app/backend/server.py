"from fastapi import FastAPI, APIRouter
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Dict
import uuid
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
from io import BytesIO


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix=\"/api\")


# Define Models
class ExpenseItem(BaseModel):
    category: str
    subcategory: str
    amount: float
    frequency: str = \"monthly\"  # monthly, bimonthly

class ExpenseData(BaseModel):
    monthly_income: float
    expenses: List[ExpenseItem]

class SavingsProjection(BaseModel):
    monthly_income: float
    total_expenses: float
    monthly_savings: float
    target_amount: float
    years: int
    required_monthly_savings: float
    shortfall: float
    expenses_by_category: Dict[str, float]


# Add your routes to the router instead of directly to app
@api_router.get(\"/\")
async def root():
    return {\"message\": \"Hello World\"}


@api_router.post(\"/calculate-projection\")
async def calculate_projection(data: ExpenseData) -> SavingsProjection:
    \"\"\"Calculate savings projection based on expenses\"\"\"
    # Calculate total monthly expenses
    total_expenses = 0
    for expense in data.expenses:
        if expense.frequency == \"bimonthly\":
            total_expenses += expense.amount / 2
        else:
            total_expenses += expense.amount
    
    # Calculate savings
    monthly_savings = data.monthly_income - total_expenses
    
    # Target: 1 Crore in 5 years
    target_amount = 10000000  # 1 Crore
    years = 5
    months = years * 12
    
    # Required monthly savings
    required_monthly_savings = target_amount / months
    
    # Shortfall
    shortfall = required_monthly_savings - monthly_savings
    
    # Group expenses by category
    expenses_by_category = {}
    for expense in data.expenses:
        amount = expense.amount if expense.frequency == \"monthly\" else expense.amount / 2
        if expense.category in expenses_by_category:
            expenses_by_category[expense.category] += amount
        else:
            expenses_by_category[expense.category] = amount
    
    return SavingsProjection(
        monthly_income=data.monthly_income,
        total_expenses=total_expenses,
        monthly_savings=monthly_savings,
        target_amount=target_amount,
        years=years,
        required_monthly_savings=required_monthly_savings,
        shortfall=shortfall,
        expenses_by_category=expenses_by_category
    )


@api_router.post(\"/export-excel\")
async def export_excel(data: ExpenseData):
    \"\"\"Generate and return Excel file with expense tracking and projections\"\"\"
    
    # Calculate projections
    projection = await calculate_projection(data)
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Monthly Expenses
    ws1 = wb.active
    ws1.title = \"Monthly Expenses\"
    
    # Header styling
    header_fill = PatternFill(start_color=\"4472C4\", end_color=\"4472C4\", fill_type=\"solid\")
    header_font = Font(color=\"FFFFFF\", bold=True, size=12)
    
    # Add title
    ws1['A1'] = \"MONTHLY EXPENDITURE TRACKER\"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f\"Monthly Income: ₹{data.monthly_income:,.2f}\"
    ws1['A2'].font = Font(bold=True, size=11)
    ws1.merge_cells('A2:D2')
    
    # Expense headers
    ws1['A4'] = \"Category\"
    ws1['B4'] = \"Subcategory\"
    ws1['C4'] = \"Amount (₹)\"
    ws1['D4'] = \"Frequency\"
    
    for cell in ['A4', 'B4', 'C4', 'D4']:
        ws1[cell].fill = header_fill
        ws1[cell].font = header_font
        ws1[cell].alignment = Alignment(horizontal='center')
    
    # Add expenses
    row = 5
    for expense in data.expenses:
        ws1[f'A{row}'] = expense.category
        ws1[f'B{row}'] = expense.subcategory
        ws1[f'C{row}'] = expense.amount
        ws1[f'D{row}'] = expense.frequency
        row += 1
    
    # Summary
    row += 1
    ws1[f'A{row}'] = \"TOTAL MONTHLY EXPENSES\"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'C{row}'] = projection.total_expenses
    ws1[f'C{row}'].font = Font(bold=True)
    
    row += 1
    ws1[f'A{row}'] = \"MONTHLY SAVINGS\"
    ws1[f'A{row}'].font = Font(bold=True, color=\"008000\" if projection.monthly_savings > 0 else \"FF0000\")
    ws1[f'C{row}'] = projection.monthly_savings
    ws1[f'C{row}'].font = Font(bold=True, color=\"008000\" if projection.monthly_savings > 0 else \"FF0000\")
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    
    # Sheet 2: Category Breakdown
    ws2 = wb.create_sheet(\"Category Breakdown\")
    ws2['A1'] = \"EXPENSE BREAKDOWN BY CATEGORY\"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:C1')
    
    ws2['A3'] = \"Category\"
    ws2['B3'] = \"Amount (₹)\"
    ws2['C3'] = \"Percentage\"
    
    for cell in ['A3', 'B3', 'C3']:
        ws2[cell].fill = header_fill
        ws2[cell].font = header_font
        ws2[cell].alignment = Alignment(horizontal='center')
    
    row = 4
    for category, amount in projection.expenses_by_category.items():
        percentage = (amount / projection.total_expenses * 100) if projection.total_expenses > 0 else 0
        ws2[f'A{row}'] = category
        ws2[f'B{row}'] = amount
        ws2[f'C{row}'] = f\"{percentage:.1f}%\"
        row += 1
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    
    # Sheet 3: 5-Year Projection
    ws3 = wb.create_sheet(\"5-Year Projection\")
    ws3['A1'] = \"PATH TO ₹1 CRORE (5 YEARS)\"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:D1')
    
    ws3['A3'] = \"Target Amount\"
    ws3['B3'] = f\"₹{projection.target_amount:,.2f}\"
    ws3['B3'].font = Font(bold=True)
    
    ws3['A4'] = \"Time Period\"
    ws3['B4'] = f\"{projection.years} years (60 months)\"
    
    ws3['A5'] = \"Current Monthly Savings\"
    ws3['B5'] = f\"₹{projection.monthly_savings:,.2f}\"
    ws3['B5'].font = Font(color=\"008000\" if projection.monthly_savings > 0 else \"FF0000\")
    
    ws3['A6'] = \"Required Monthly Savings\"
    ws3['B6'] = f\"₹{projection.required_monthly_savings:,.2f}\"
    ws3['B6'].font = Font(bold=True)
    
    ws3['A7'] = \"Monthly Shortfall/Surplus\"
    shortfall_amount = projection.shortfall
    ws3['B7'] = f\"₹{abs(shortfall_amount):,.2f} {'(Shortfall)' if shortfall_amount > 0 else '(Surplus)'}\"
    ws3['B7'].font = Font(color=\"FF0000\" if shortfall_amount > 0 else \"008000\", bold=True)
    
    ws3['A9'] = \"Month-by-Month Projection (Assuming Current Savings Rate)\"
    ws3['A9'].font = Font(bold=True)
    ws3.merge_cells('A9:D9')
    
    ws3['A11'] = \"Month\"
    ws3['B11'] = \"Monthly Savings\"
    ws3['C11'] = \"Cumulative Savings\"
    ws3['D11'] = \"Target Progress\"
    
    for cell in ['A11', 'B11', 'C11', 'D11']:
        ws3[cell].fill = header_fill
        ws3[cell].font = header_font
        ws3[cell].alignment = Alignment(horizontal='center')
    
    # Generate 60 months projection
    cumulative = 0
    for month in range(1, 61):
        row = 11 + month
        cumulative += projection.monthly_savings
        progress = (cumulative / projection.target_amount * 100) if projection.target_amount > 0 else 0
        
        ws3[f'A{row}'] = month
        ws3[f'B{row}'] = projection.monthly_savings
        ws3[f'C{row}'] = cumulative
        ws3[f'D{row}'] = f\"{progress:.1f}%\"
    
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 18
    
    # Sheet 4: Recommendations
    ws4 = wb.create_sheet(\"Recommendations\")
    ws4['A1'] = \"RECOMMENDATIONS TO REACH YOUR GOAL\"
    ws4['A1'].font = Font(bold=True, size=14, color=\"4472C4\")
    ws4.merge_cells('A1:C1')
    
    row = 3
    if projection.shortfall > 0:
        ws4[f'A{row}'] = \"Current Situation:\"
        ws4[f'A{row}'].font = Font(bold=True)
        row += 1
        ws4[f'A{row}'] = f\"You need to save ₹{projection.required_monthly_savings:,.2f}/month to reach ₹1 Crore in 5 years.\"
        row += 1
        ws4[f'A{row}'] = f\"Currently, you're saving ₹{projection.monthly_savings:,.2f}/month.\"
        row += 1
        ws4[f'A{row}'] = f\"You need to increase your savings by ₹{projection.shortfall:,.2f}/month.\"
        ws4[f'A{row}'].font = Font(bold=True, color=\"FF0000\")
        row += 2
        
        ws4[f'A{row}'] = \"Recommended Actions:\"
        ws4[f'A{row}'].font = Font(bold=True, underline=\"single\")
        row += 1
        
        # Find top 3 expense categories
        sorted_categories = sorted(projection.expenses_by_category.items(), key=lambda x: x[1], reverse=True)[:3]
        ws4[f'A{row}'] = f\"1. Review your top expense categories:\"
        row += 1
        for i, (cat, amt) in enumerate(sorted_categories, 1):
            ws4[f'A{row}'] = f\"   • {cat}: ₹{amt:,.2f}/month\"
            row += 1
        
        row += 1
        ws4[f'A{row}'] = \"2. Consider increasing your income through:\"
        row += 1
        ws4[f'A{row}'] = \"   • Side projects or freelancing\"
        row += 1
        ws4[f'A{row}'] = \"   • Skill development for promotions\"
        row += 1
        ws4[f'A{row}'] = \"   • Investment returns\"
        
    else:
        ws4[f'A{row}'] = \"Congratulations! 🎉\"
        ws4[f'A{row}'].font = Font(bold=True, color=\"008000\", size=12)
        row += 1
        ws4[f'A{row}'] = f\"You're saving ₹{projection.monthly_savings:,.2f}/month, which is more than required!\"
        row += 1
        ws4[f'A{row}'] = \"Keep up the great work and stay consistent with your savings.\"
    
    ws4.column_dimensions['A'].width = 80
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\",
        headers={\"Content-Disposition\": \"attachment; filename=expense_tracker.xlsx\"}
    )


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=[\"*\"],
    allow_headers=[\"*\"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event(\"shutdown\")
async def shutdown_db_client():
    client.close()
"
